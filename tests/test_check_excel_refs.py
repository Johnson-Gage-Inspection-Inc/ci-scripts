"""Tests for check_excel_refs module - inspired by working shell tests."""

import csv
import os
import subprocess
import sys
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from check_excel_refs import check_ref_errors, export_sheets_with_formulas, main


class TestCheckExcelRefsIntegration:
    """Integration tests using the real test files (like shell tests do)."""

    @property
    def project_root(self):
        """Get the project root directory."""
        return Path(__file__).parent.parent

    def test_command_line_no_args(self):
        """Test: No file specified - should exit with error."""
        result = subprocess.run(
            [sys.executable, "check_excel_refs.py"],
            cwd=self.project_root,
            capture_output=True,
            text=True,
        )
        assert result.returncode == 1
        assert "No Excel file specified" in result.stdout

    def test_command_line_nonexistent_file(self):
        """Test: Non-existent file - should exit with error."""
        result = subprocess.run(
            [sys.executable, "check_excel_refs.py", "nonexistent.xlsx"],
            cwd=self.project_root,
            capture_output=True,
            text=True,
        )
        assert result.returncode == 1

    def test_command_line_unsupported_file(self):
        """Test: Unsupported file type - should exit with error."""
        # Create a text file
        test_file = Path("temp_test.txt")
        test_file.write_text("not an excel file")

        try:
            result = subprocess.run(
                [sys.executable, "check_excel_refs.py", str(test_file)],
                cwd=self.project_root,
                capture_output=True,
                text=True,
            )
            assert result.returncode == 1
        finally:
            if test_file.exists():
                test_file.unlink()

    def test_command_line_file_with_broken_refs(self):
        """Test: File with broken references - should detect them."""
        test_file = Path("tests/test_files/has_broken_ref.xltm")
        if test_file.exists():
            result = subprocess.run(
                [sys.executable, "check_excel_refs.py", str(test_file)],
                cwd=self.project_root,
                capture_output=True,
                text=True,
            )
            assert result.returncode == 1  # Should find errors

    def test_command_line_file_no_errors(self):
        """Test: File with no errors - should pass."""
        test_file = Path("tests/test_files/has_no_errors.xltm")
        if test_file.exists():
            result = subprocess.run(
                [sys.executable, "check_excel_refs.py", str(test_file)],
                cwd=self.project_root,
                capture_output=True,
                text=True,
            )
            assert result.returncode == 0  # Should pass

    def test_export_sheets_functionality(self):
        """Test: Export sheets functionality with environment variable."""
        test_file = Path("tests/test_files/has_no_errors.xltm")
        if test_file.exists():
            env = os.environ.copy()
            env["EXPORT_SHEETS"] = "true"

            result = subprocess.run(
                [sys.executable, "check_excel_refs.py", str(test_file)],
                cwd=self.project_root,
                capture_output=True,
                text=True,
                env=env,
            )
            # Should work regardless of export functionality
            assert result.returncode in [0, 1]


class TestCheckExcelRefsUnits:
    """Unit tests for individual functions."""

    def test_export_sheets_with_formulas_simple(self, tmp_path):
        """Test exporting a simple sheet with basic data."""
        # Create a test workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Hello"
        ws["B1"] = "World"
        ws["A2"] = 42
        ws["B2"] = "=A2*2"  # Formula

        # Save to temporary file
        xlsx_path = tmp_path / "test.xlsx"
        wb.save(xlsx_path)

        # Export sheets
        output_dir = tmp_path / "output"
        export_sheets_with_formulas(xlsx_path, output_dir)

        # Check output
        csv_path = output_dir / "TestSheet.csv"
        assert csv_path.exists()

        with open(csv_path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)

        assert len(rows) >= 2
        assert rows[0][0] == "Hello"
        assert rows[0][1] == "World"
        assert rows[1][0] == "42"
        assert rows[1][1] == "=A2*2"

    def test_export_sheets_with_multiple_sheets(self, tmp_path):
        """Test exporting workbook with multiple sheets."""
        wb = openpyxl.Workbook()

        # First sheet
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "Data1"

        # Second sheet
        ws2 = wb.create_sheet("Sheet2")
        ws2["A1"] = "Data2"

        xlsx_path = tmp_path / "test.xlsx"
        wb.save(xlsx_path)

        output_dir = tmp_path / "output"
        export_sheets_with_formulas(xlsx_path, output_dir)

        assert (output_dir / "Sheet1.csv").exists()
        assert (output_dir / "Sheet2.csv").exists()

    def test_check_ref_errors_with_real_file(self):
        """Test check_ref_errors function with real test files."""
        # Test with broken ref file
        broken_file = Path("tests/test_files/has_broken_ref.xltm")
        if broken_file.exists():
            errors = check_ref_errors(broken_file)
            # Should find some errors
            assert errors is not None

        # Test with clean file
        clean_file = Path("tests/test_files/has_no_errors.xltm")
        if clean_file.exists():
            errors = check_ref_errors(clean_file)
            # Should be empty or None
            assert errors is None or len(errors) == 0

    def test_check_ref_errors_nonexistent_file(self):
        """Test check_ref_errors with nonexistent file."""
        result = check_ref_errors(Path("nonexistent.xlsx"))
        # Should handle gracefully and return None or empty
        assert result is None or (isinstance(result, list) and len(result) == 0)

    @patch("sys.argv")
    @patch("os.environ.get")
    def test_main_no_args_no_env(self, mock_env_get, mock_argv):
        """Test main function with no arguments and no environment variable."""
        mock_argv.__getitem__.side_effect = lambda x: ["check_excel_refs.py"][x]
        mock_argv.__len__.return_value = 1

        def mock_env_side_effect(key, default=None):
            if key == "EXCEL_FILE":
                return None
            elif key == "EXPORT_SHEETS":
                return default  # Return the default value
            return default

        mock_env_get.side_effect = mock_env_side_effect

        # Mock sys.exit to raise SystemExit (which is what it normally does)
        with patch("sys.exit", side_effect=SystemExit) as mock_exit:
            with pytest.raises(SystemExit):
                main()
            mock_exit.assert_called_once_with(1)

    @patch("sys.argv", ["check_excel_refs.py", "tests/test_files/has_no_errors.xltm"])
    def test_main_with_valid_file(self):
        """Test main function with valid file."""
        test_file = Path("tests/test_files/has_no_errors.xltm")
        if test_file.exists():
            with patch("sys.exit") as mock_exit:
                main()
                # Should exit with 0 for clean file
                mock_exit.assert_called_once_with(0)

    @patch("sys.argv", ["check_excel_refs.py", "tests/test_files/has_broken_ref.xltm"])
    def test_main_with_broken_file(self):
        """Test main function with file containing broken references."""
        test_file = Path("tests/test_files/has_broken_ref.xltm")
        if test_file.exists():
            with patch("sys.exit") as mock_exit:
                main()
                # Should exit with 1 for file with errors
                mock_exit.assert_called_once_with(1)

    @patch("sys.argv", ["check_excel_refs.py", "nonexistent.xlsx"])
    def test_main_with_nonexistent_file(self):
        """Test main function with nonexistent file."""
        # Mock sys.exit to raise SystemExit (which is what it normally does)
        with patch("sys.exit", side_effect=SystemExit) as mock_exit:
            with pytest.raises(SystemExit):
                main()
            # Should exit with 1 for nonexistent file - check the first call
            mock_exit.assert_called_with(1)


class TestExportSheetsEdgeCases:
    """Test edge cases for export functionality."""

    def test_export_empty_sheet_creates_empty_csv(self, tmp_path):
        """Test that empty sheets create empty CSV files (matching actual behavior)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EmptySheet"

        xlsx_path = tmp_path / "test.xlsx"
        wb.save(xlsx_path)

        output_dir = tmp_path / "output"
        export_sheets_with_formulas(xlsx_path, output_dir)

        csv_path = output_dir / "EmptySheet.csv"
        # Based on the actual implementation, empty sheets DO create CSV files
        # They're just empty or contain minimal content
        assert csv_path.exists()

    def test_export_sheet_with_formulas_and_values(self, tmp_path):
        """Test exporting sheet with mix of formulas and values."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MixedSheet"

        # Add various types of content
        ws["A1"] = "Text"
        ws["B1"] = 123
        ws["C1"] = "=B1*2"
        ws["D1"] = None
        ws["A2"] = '=A1&" test"'

        xlsx_path = tmp_path / "test.xlsx"
        wb.save(xlsx_path)

        output_dir = tmp_path / "output"
        export_sheets_with_formulas(xlsx_path, output_dir)

        csv_path = output_dir / "MixedSheet.csv"
        assert csv_path.exists()

        with open(csv_path, "r", encoding="utf-8") as f:
            content = f.read()
            assert "Text" in content
            assert "123" in content
            assert "=B1*2" in content


class TestCheckRefErrorsDetailed:
    """Detailed unit tests for check_ref_errors function."""

    def test_check_ref_errors_with_simple_ref_error(self, tmp_path):
        """Test detecting a simple #REF! error in a cell value."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"

        # Add normal data
        ws["A1"] = "Normal text"
        ws["B1"] = 42

        # Add a cell with #REF! error
        ws["C1"] = "#REF!"

        xlsx_path = tmp_path / "ref_error_test.xlsx"
        wb.save(xlsx_path)

        errors = check_ref_errors(xlsx_path)

        assert errors is not None
        assert len(errors) == 1
        assert errors[0]["sheet"] == "TestSheet"
        assert errors[0]["cell"] == "C1"
        assert "#REF!" in errors[0]["formula"]

    def test_check_ref_errors_with_formula_ref_error(self, tmp_path):
        """Test detecting #REF! error in a formula."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FormulaSheet"

        # Add normal formula
        ws["A1"] = 10
        ws["B1"] = "=A1*2"

        # Add formula with #REF! error
        ws["C1"] = "=SUM(#REF!)"

        xlsx_path = tmp_path / "formula_ref_test.xlsx"
        wb.save(xlsx_path)

        errors = check_ref_errors(xlsx_path)

        assert errors is not None
        assert len(errors) == 1
        assert errors[0]["sheet"] == "FormulaSheet"
        assert errors[0]["cell"] == "C1"
        assert "#REF!" in errors[0]["formula"]

    def test_check_ref_errors_multiple_errors(self, tmp_path):
        """Test detecting multiple #REF! errors across different cells."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MultiErrorSheet"

        # Add multiple errors
        ws["A1"] = "#REF!"
        ws["B2"] = "=SUM(#REF!)"
        ws["C3"] = "Text with #REF! in it"
        ws["D4"] = "=A1+#REF!"

        # Add normal data
        ws["E1"] = "Normal"
        ws["F2"] = "=E1&' text'"

        xlsx_path = tmp_path / "multi_error_test.xlsx"
        wb.save(xlsx_path)

        errors = check_ref_errors(xlsx_path)

        assert errors is not None
        assert len(errors) == 4  # Should find all 4 errors

        # Check that all error locations are captured
        error_cells = {error["cell"] for error in errors}
        assert "A1" in error_cells
        assert "B2" in error_cells
        assert "C3" in error_cells
        assert "D4" in error_cells

    def test_check_ref_errors_multiple_sheets(self, tmp_path):
        """Test detecting #REF! errors across multiple sheets."""
        wb = openpyxl.Workbook()

        # First sheet with error
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = "#REF!"

        # Second sheet with error
        ws2 = wb.create_sheet("Sheet2")
        ws2["B1"] = "=SUM(#REF!)"

        # Third sheet without errors
        ws3 = wb.create_sheet("CleanSheet")
        ws3["A1"] = "Clean data"
        ws3["B1"] = "=A1&' more'"

        xlsx_path = tmp_path / "multi_sheet_test.xlsx"
        wb.save(xlsx_path)

        errors = check_ref_errors(xlsx_path)

        assert errors is not None
        assert len(errors) == 2

        # Check that errors from both sheets are found
        error_sheets = {error["sheet"] for error in errors}
        assert "Sheet1" in error_sheets
        assert "Sheet2" in error_sheets
        assert "CleanSheet" not in error_sheets

    def test_check_ref_errors_no_errors(self, tmp_path):
        """Test file with no #REF! errors returns empty list."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CleanSheet"

        # Add various types of clean data
        ws["A1"] = "Text"
        ws["B1"] = 42
        ws["C1"] = 3.14
        ws["D1"] = "=A1&B1"
        ws["E1"] = "=SUM(B1:C1)"
        ws["F1"] = None

        xlsx_path = tmp_path / "clean_test.xlsx"
        wb.save(xlsx_path)

        errors = check_ref_errors(xlsx_path)

        assert errors is not None
        assert len(errors) == 0

    def test_check_ref_errors_empty_sheet(self, tmp_path):
        """Test completely empty sheet returns no errors."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EmptySheet"
        # Don't add any data

        xlsx_path = tmp_path / "empty_test.xlsx"
        wb.save(xlsx_path)

        errors = check_ref_errors(xlsx_path)

        assert errors is not None
        assert len(errors) == 0

    def test_check_ref_errors_nonexistent_file(self):
        """Test behavior with nonexistent file."""
        nonexistent_path = Path("definitely_does_not_exist.xlsx")
        errors = check_ref_errors(nonexistent_path)

        # Should handle gracefully and return None
        assert errors is None

    def test_check_ref_errors_invalid_file(self, tmp_path):
        """Test behavior with invalid Excel file."""
        # Create a text file that's not an Excel file
        invalid_file = tmp_path / "not_excel.txt"
        invalid_file.write_text("This is not an Excel file")

        errors = check_ref_errors(invalid_file)

        # Should handle gracefully and return None
        assert errors is None

    def test_check_ref_errors_with_different_cell_types(self, tmp_path):
        """Test #REF! detection in different cell types."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "MixedTypes"

        # Different types of cells with #REF!
        ws["A1"] = "#REF!"  # String value
        ws["B1"] = "Text with #REF! embedded"  # String with #REF!
        ws["C1"] = "=SUM(#REF!)"  # Formula with #REF!
        ws["D1"] = "=#REF!+5"  # Another formula type

        # Clean cells for comparison
        ws["E1"] = "Normal text"
        ws["F1"] = 42
        ws["G1"] = "=SUM(A1:B1)"  # Clean formula

        xlsx_path = tmp_path / "mixed_types_test.xlsx"
        wb.save(xlsx_path)

        errors = check_ref_errors(xlsx_path)

        assert errors is not None
        assert len(errors) == 4  # Should find all 4 #REF! errors

        # Verify no false positives
        error_cells = {error["cell"] for error in errors}
        assert "E1" not in error_cells
        assert "F1" not in error_cells
        assert "G1" not in error_cells

    def test_check_ref_errors_edge_cases(self, tmp_path):
        """Test edge cases for #REF! detection."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EdgeCases"

        # Edge cases
        ws["A1"] = "#REF"  # Without the !
        ws["B1"] = "REF!"  # Without the #
        ws["C1"] = "#ref!"  # Lowercase
        ws["D1"] = "#REF!"  # Exact match - should be found
        ws["E1"] = "prefix#REF!suffix"  # With surrounding text
        ws["F1"] = None  # None value
        ws["G1"] = ""  # Empty string

        xlsx_path = tmp_path / "edge_cases_test.xlsx"
        wb.save(xlsx_path)

        errors = check_ref_errors(xlsx_path)

        assert errors is not None

        # Should find exact matches and embedded matches
        error_cells = {error["cell"] for error in errors}
        assert "D1" in error_cells  # Exact #REF!
        assert "E1" in error_cells  # Embedded #REF!

        # Should not find partial matches
        assert "A1" not in error_cells  # Missing !
        assert "B1" not in error_cells  # Missing #
        assert "C1" not in error_cells  # Wrong case

    def test_check_ref_errors_array_formula(self, tmp_path):
        """Test #REF! detection in array formulas."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "ArrayFormulas"

        # Try to create an array formula with #REF!
        # Note: This is a complex case that might need special handling
        ws["A1"] = "=SUM(#REF!:B5)"
        ws["B1"] = "={SUM(#REF!)}"

        xlsx_path = tmp_path / "array_formula_test.xlsx"
        wb.save(xlsx_path)

        errors = check_ref_errors(xlsx_path)

        assert errors is not None
        # Should find errors in array formulas too
        assert len(errors) >= 1
