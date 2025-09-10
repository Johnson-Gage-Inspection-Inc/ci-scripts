"""Integration tests against the xl-test repository.

This module contains integration tests that create actual Pull Requests
in the xl-test repository to verify that the CI scripts work end-to-end
in a real GitHub Actions environment.

Tests follow this approach:
1. Check out a new branch in xl-test
2. Update Book1.xltx using openpyxl (either valid changes or broken refs)
3. Commit and push changes
4. Create a pull request
5. Assert that the checks pass/fail as expected
6. Clean up (merge valid PRs, close broken ones)
"""

import base64
import os
import sys
import time
from datetime import datetime
from typing import Dict, List

import pytest
import requests
from openpyxl import Workbook


class GitHubIntegrationError(Exception):
    """Custom exception for GitHub integration test failures."""

    pass


class XLTestIntegration:
    """Helper class for integration testing with xl-test repository."""

    def __init__(self):
        self.repo_owner = "Johnson-Gage-Inspection-Inc"
        self.repo_name = "xl-test"
        self.api_base = "https://api.github.com"
        self.token = os.environ.get("GITHUB_TOKEN")

        if not self.token:
            raise GitHubIntegrationError(
                "GITHUB_TOKEN environment variable is required for integration tests"
            )

    @property
    def headers(self) -> Dict[str, str]:
        """HTTP headers for GitHub API requests."""
        return {
            "Authorization": f"token {self.token}",
            "Accept": "application/vnd.github.v3+json",
            "User-Agent": "ci-scripts-integration-test",
        }

    def get_file_content(self, file_path: str, branch: str = "main") -> Dict:
        """Get the content and metadata of a file from the repository."""
        url = f"{self.api_base}/repos/{self.repo_owner}/{self.repo_name}/contents/{file_path}"
        params = {"ref": branch}
        response = requests.get(url, headers=self.headers, params=params)
        response.raise_for_status()
        return response.json() or {}

    def create_test_branch(self, branch_name: str, base_branch: str = "main") -> str:
        """Create a test branch in the xl-test repository."""
        # Get the SHA of the base branch
        url = f"{self.api_base}/repos/{self.repo_owner}/{self.repo_name}/git/refs/heads/{base_branch}"
        response = requests.get(url, headers=self.headers)
        response.raise_for_status()

        base_sha = response.json()["object"]["sha"]

        # Create new branch
        url = f"{self.api_base}/repos/{self.repo_owner}/{self.repo_name}/git/refs"
        data = {"ref": f"refs/heads/{branch_name}", "sha": base_sha}
        response = requests.post(url, json=data, headers=self.headers)
        response.raise_for_status()

        return str(base_sha)

    def create_valid_excel_file(self) -> bytes:
        """Create a valid Excel file with proper structure."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Add some valid data
        ws["A1"] = "Test Data"
        ws["B1"] = "Value"
        ws["A2"] = "Item 1"
        ws["B2"] = 100
        ws["A3"] = "Item 2"
        ws["B3"] = 200

        # Add a valid formula (no #REF! errors)
        ws["C2"] = "=B2*2"
        ws["C3"] = "=B3*2"
        ws["C4"] = "=SUM(C2:C3)"

        # Save to bytes
        from io import BytesIO

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def create_broken_excel_file(self) -> bytes:
        """Create an Excel file with #REF! errors."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Add some data
        ws["A1"] = "Test Data"
        ws["B1"] = "Value"
        ws["A2"] = "Item 1"
        ws["B2"] = 100

        # Add formulas that will create #REF! errors
        # These formulas reference cells that don't exist or will be deleted
        ws["C2"] = "=ZZ99999"  # Invalid cell reference
        ws["C3"] = "=#REF!"  # Direct #REF! error
        ws["C4"] = "=SUM(#REF!:C3)"  # Formula with #REF!

        # Save to bytes
        from io import BytesIO

        buffer = BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

    def update_file_in_repo(
        self, file_path: str, content: bytes, branch_name: str, message: str
    ):
        """Update a file in the repository."""
        # Get current file to get its SHA
        try:
            file_info = self.get_file_content(file_path, branch_name)
            current_sha = file_info["sha"]
        except requests.exceptions.HTTPError:
            # File doesn't exist, so no SHA needed
            current_sha = None

        # Encode content to base64
        content_b64 = base64.b64encode(content).decode("utf-8")

        # Update file
        url = f"{self.api_base}/repos/{self.repo_owner}/{self.repo_name}/contents/{file_path}"
        data = {"message": message, "content": content_b64, "branch": branch_name}
        if current_sha:
            data["sha"] = current_sha

        response = requests.put(url, json=data, headers=self.headers)
        response.raise_for_status()
        return response.json()

    def create_pull_request(
        self, branch_name: str, title: str, body: str, base_branch: str = "main"
    ) -> int:
        """Create a pull request."""
        url = f"{self.api_base}/repos/{self.repo_owner}/{self.repo_name}/pulls"
        data = {"title": title, "body": body, "head": branch_name, "base": base_branch}
        response = requests.post(url, json=data, headers=self.headers)
        response.raise_for_status()

        return int(response.json()["number"])

    def get_pull_request_status(self, pr_number: int) -> Dict:
        """Get the current status of a pull request."""
        url = f"{self.api_base}/repos/{self.repo_owner}/{self.repo_name}/pulls/{pr_number}"
        response = requests.get(url, headers=self.headers)
        response.raise_for_status()

        pr_data = response.json()

        # Get status checks
        sha = pr_data["head"]["sha"]

        # Get both status (legacy) and check runs (modern)
        status_url = f"{self.api_base}/repos/{self.repo_owner}/{self.repo_name}/commits/{sha}/status"
        status_response = requests.get(status_url, headers=self.headers)
        status_response.raise_for_status()

        checks_url = f"{self.api_base}/repos/{self.repo_owner}/{self.repo_name}/commits/{sha}/check-runs"
        checks_response = requests.get(checks_url, headers=self.headers)
        checks_response.raise_for_status()

        return {
            "pr": pr_data,
            "status": status_response.json(),
            "checks": checks_response.json(),
            "sha": sha,
        }

    def wait_for_checks(self, pr_number: int, timeout: int = 600) -> bool:
        """Wait for CI checks to complete on a pull request."""
        start_time = time.time()
        print(f"Waiting for CI checks on PR #{pr_number}...")

        # Start with shorter polling interval for faster response
        poll_interval = 15  # Start with 15s, increase to 30s later
        checks_started = False

        while time.time() - start_time < timeout:
            elapsed = int(time.time() - start_time)
            status_info = self.get_pull_request_status(pr_number)
            pr_data = status_info["pr"]

            # Check if PR has merge conflicts
            if not pr_data.get("mergeable", True):
                print(f"PR #{pr_number} has merge conflicts, cannot run CI checks")
                print("Treating as a failed test scenario")
                return False

            # Check if PR is closed/merged (might have been handled externally)
            if pr_data["state"] == "closed":
                if pr_data.get("merged", False):
                    print(f"PR #{pr_number} was merged externally")
                    return True
                else:
                    print(f"PR #{pr_number} was closed externally")
                    return False

            # Check legacy status API
            status_state = status_info["status"]["state"]

            # Check modern check runs API
            check_runs = status_info["checks"]["check_runs"]

            print(
                f"[{elapsed}s] Status: {status_state}, Check runs: {len(check_runs)}, Mergeable: {pr_data.get('mergeable', 'unknown')}"
            )

            # If we have check runs, check their status
            if check_runs:
                checks_started = True
                all_completed = all(run["status"] == "completed" for run in check_runs)

                # Show progress details
                running_checks = [
                    run["name"] for run in check_runs if run["status"] != "completed"
                ]
                if running_checks:
                    print(f"  Running checks: {', '.join(running_checks)}")

                if all_completed:
                    success = all(run["conclusion"] == "success" for run in check_runs)
                    failed_checks = [
                        run["name"]
                        for run in check_runs
                        if run["conclusion"] != "success"
                    ]
                    if failed_checks:
                        print(f"  Failed checks: {', '.join(failed_checks)}")
                    print(f"All checks completed. Success: {success}")
                    return success

            # Fallback to legacy status
            if status_state in ["success", "failure", "error"]:
                success = status_state == "success"
                print(f"Legacy status completed. Success: {success}")
                return success

            # Adaptive polling: use longer intervals once checks are running
            if checks_started:
                poll_interval = 30

            print(f"Checks still running, waiting {poll_interval} seconds...")
            time.sleep(poll_interval)

        raise GitHubIntegrationError(
            f"Timeout waiting for CI checks on PR #{pr_number}"
        )

    def merge_pull_request(self, pr_number: int, merge_method: str = "merge"):
        """Merge a pull request."""
        url = f"{self.api_base}/repos/{self.repo_owner}/{self.repo_name}/pulls/{pr_number}/merge"
        data = {"merge_method": merge_method}
        response = requests.put(url, json=data, headers=self.headers)
        response.raise_for_status()

    def close_pull_request(self, pr_number: int):
        """Close a pull request."""
        url = f"{self.api_base}/repos/{self.repo_owner}/{self.repo_name}/pulls/{pr_number}"
        data = {"state": "closed"}
        response = requests.patch(url, json=data, headers=self.headers)
        response.raise_for_status()

    def delete_branch(self, branch_name: str):
        """Delete a test branch."""
        url = f"{self.api_base}/repos/{self.repo_owner}/{self.repo_name}/git/refs/heads/{branch_name}"
        response = requests.delete(url, headers=self.headers)
        response.raise_for_status()


@pytest.mark.integration
@pytest.mark.skipif(
    not os.environ.get("GITHUB_TOKEN") or os.environ.get("SKIP_INTEGRATION") == "true",
    reason="GITHUB_TOKEN required and SKIP_INTEGRATION not set",
)
class TestXLIntegration:
    """Integration tests with xl-test repository."""

    def setup_method(self):
        """Set up test instance."""
        self.xl_test = XLTestIntegration()
        self.test_branches: List[str] = []
        self.test_prs: List[int] = []

    def teardown_method(self):
        """Clean up test resources."""
        # Close PRs and delete branches
        for pr_number in self.test_prs:
            try:
                self.xl_test.close_pull_request(pr_number)
                print(f"Closed PR #{pr_number}")
            except Exception as e:
                print(f"Warning: Failed to close PR #{pr_number}: {e}")

        for branch_name in self.test_branches:
            try:
                self.xl_test.delete_branch(branch_name)
                print(f"Deleted branch {branch_name}")
            except Exception as e:
                print(f"Warning: Failed to delete branch {branch_name}: {e}")

    def test_ci_scripts_end_to_end_workflow(self):
        """Test complete CI workflow: create both PRs in parallel, then handle sequentially."""
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        python_version = f"{sys.version_info.major}{sys.version_info.minor}"

        # Create both branches and PRs in parallel for faster execution
        broken_branch = f"ci-test-broken-py{python_version}-{timestamp}"
        valid_branch = f"ci-test-valid-py{python_version}-{timestamp}-valid"

        print("=== Creating both test branches and PRs in parallel ===")

        try:
            # Phase 1: Create both branches and PRs simultaneously
            print(f"Creating branches: {broken_branch} and {valid_branch}")

            # Create both branches
            self.xl_test.create_test_branch(broken_branch)
            self.test_branches.append(broken_branch)

            self.xl_test.create_test_branch(valid_branch)
            self.test_branches.append(valid_branch)

            # Update both files with their respective content
            broken_excel_content = self.xl_test.create_broken_excel_file()
            valid_excel_content = self.xl_test.create_valid_excel_file()

            # Update broken branch
            self.xl_test.update_file_in_repo(
                "Book1.xltx",
                broken_excel_content,
                broken_branch,
                f"test: Update Book1.xltx with broken references - {timestamp}",
            )
            print("Updated broken branch with #REF! errors")

            # Update valid branch
            self.xl_test.update_file_in_repo(
                "Book1.xltx",
                valid_excel_content,
                valid_branch,
                f"test: Update Book1.xltx with valid changes - {timestamp}",
            )
            print("Updated valid branch with clean content")

            # Create both PRs
            broken_pr_number = self.xl_test.create_pull_request(
                broken_branch,
                f"[CI Test] Broken reference detection - {timestamp}",
                f"Integration test to verify CI scripts can detect broken Excel references.\n\nBranch: {broken_branch}\nTimestamp: {timestamp}\nPython: {python_version}",
            )
            self.test_prs.append(broken_pr_number)
            print(f"Created broken refs PR #{broken_pr_number}")

            valid_pr_number = self.xl_test.create_pull_request(
                valid_branch,
                f"[CI Test] Valid Excel file test - {timestamp}",
                f"Integration test to verify CI scripts work with valid Excel files.\n\nBranch: {valid_branch}\nTimestamp: {timestamp}\nPython: {python_version}",
            )
            self.test_prs.append(valid_pr_number)
            print(f"Created valid file PR #{valid_pr_number}")

            # Phase 2: Wait for broken PR to fail, then close it
            print("=== Waiting for broken PR to fail ===")
            # Use shorter timeout for broken PR since we expect it to fail quickly
            broken_success = self.xl_test.wait_for_checks(broken_pr_number, timeout=240)
            assert (
                not broken_success
            ), f"CI checks should have failed for broken PR #{broken_pr_number}"
            print(
                f"✅ CI correctly detected broken references in PR #{broken_pr_number}"
            )

            # Close the broken PR first
            self.xl_test.close_pull_request(broken_pr_number)
            print(f"✅ Closed broken refs PR #{broken_pr_number}")

            # Phase 3: Now merge the valid PR (should have clean merge)
            print("=== Processing valid PR (should merge cleanly) ===")
            # Use longer timeout for valid PR to allow for complete CI pipeline
            valid_success = self.xl_test.wait_for_checks(valid_pr_number, timeout=360)
            assert valid_success, f"CI checks failed for valid PR #{valid_pr_number}"
            print(f"✅ CI passed for valid Excel file in PR #{valid_pr_number}")

            # Merge the valid PR
            self.xl_test.merge_pull_request(valid_pr_number)
            print(f"✅ Merged valid PR #{valid_pr_number}")

            print("🎉 Parallel creation + sequential handling completed successfully!")

        except Exception as e:
            pytest.fail(f"End-to-end workflow test failed: {e}")


@pytest.mark.integration
def test_github_token_available():
    """Verify GITHUB_TOKEN works for repo-scoped calls (App installation token)."""
    if os.environ.get("SKIP_INTEGRATION", "").lower() == "true":
        pytest.skip("Integration tests disabled via SKIP_INTEGRATION")

    token = os.environ.get("GITHUB_TOKEN")
    if not token:
        pytest.skip("GITHUB_TOKEN not available - integration tests will be skipped")

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28",
    }

    repo = os.environ.get("GITHUB_REPOSITORY")  # e.g., "owner/repo"
    assert repo, "GITHUB_REPOSITORY not set in environment"

    r = requests.get(f"https://api.github.com/repos/{repo}", headers=headers)

    # If token is valid but lacks scope (e.g., from restricted context), skip instead of failing CI.
    if r.status_code in (401, 403):
        pytest.skip(
            f"GITHUB_TOKEN not usable in this context: {r.status_code} - {r.text}"
        )

    assert r.status_code == 200, f"Unexpected status: {r.status_code} - {r.text}"


if __name__ == "__main__":
    # Allow running integration tests directly
    pytest.main([__file__, "-v", "-m", "integration"])
