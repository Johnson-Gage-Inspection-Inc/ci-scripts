"""Pytest configuration and fixtures."""

from pathlib import Path
from typing import Generator

import pytest


@pytest.fixture
def temp_excel_file(tmp_path: Path) -> Generator[Path, None, None]:
    """Create a temporary Excel file for testing."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TestSheet"
    ws["A1"] = "Test Data"
    ws["B1"] = "=A1&' Modified'"

    excel_path = tmp_path / "test.xlsx"
    wb.save(excel_path)

    yield excel_path


@pytest.fixture
def temp_excel_with_ref_errors(tmp_path: Path) -> Generator[Path, None, None]:
    """Create a temporary Excel file with #REF! errors for testing."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ErrorSheet"
    ws["A1"] = "Normal Data"
    ws["B1"] = "=#REF!"
    ws["C1"] = "=SUM(#REF!)"

    excel_path = tmp_path / "test_with_errors.xlsx"
    wb.save(excel_path)

    yield excel_path


@pytest.fixture
def sample_csv_data(tmp_path: Path) -> Generator[Path, None, None]:
    """Create a sample CSV file for testing."""
    csv_content = """Name,Age,City
John Doe,30,New York
Jane Smith,25,Los Angeles
Bob Johnson,35,Chicago"""

    csv_path = tmp_path / "sample.csv"
    csv_path.write_text(csv_content)

    yield csv_path


@pytest.fixture(autouse=True)
def setup_test_environment(tmp_path: Path, monkeypatch) -> None:
    """Set up a clean test environment for each test."""
    # Change to temp directory to avoid polluting the workspace
    monkeypatch.chdir(tmp_path)

    # Ensure no leftover temp files
    temp_dirs = ["tmp_unzipped", "exploded", "output", "extracted"]
    for temp_dir in temp_dirs:
        temp_path = tmp_path / temp_dir
        if temp_path.exists():
            import shutil

            shutil.rmtree(temp_path)
