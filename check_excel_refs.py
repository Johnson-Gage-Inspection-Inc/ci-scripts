#!/usr/bin/env python3
"""
Check for #REF! errors in Excel files and export sheets with formulas.
"""
import csv
import os
import sys
import zipfile
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.formula import ArrayFormula
from openpyxl.worksheet.worksheet import Worksheet


def export_sheets_with_formulas(xlsx_path: Path, output_dir: Path):
    """Export all sheets from Excel file to CSV, preserving formulas."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=False, keep_links=False)
    output_dir.mkdir(parents=True, exist_ok=True)

    for sheet in wb.worksheets:
        # Skip if this is not a worksheet (e.g., chartsheet)
        if not isinstance(sheet, Worksheet):
            continue

        def get_formula_or_value(c) -> str:
            """Get the formula or value from a cell, handling all types."""
            # Handle MergedCell case
            if not hasattr(c, "value"):
                return ""

            val = c.value

            # Handle different cell types
            if hasattr(c, "data_type") and c.data_type == "f":  # Formula
                if isinstance(val, ArrayFormula):
                    return val.text if val.text else ""
                return str(val) if val is not None else ""
            elif val is None:
                return ""
            else:
                return str(val)

        csv_path = output_dir / f"{sheet.title}.csv"
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            rows = [
                [get_formula_or_value(cell) for cell in row]
                for row in sheet.iter_rows(values_only=False)
            ]
            writer.writerows(rows)

        # After writing, check if all rows are just commas
        with open(csv_path, "r+", encoding="utf-8") as f:
            content = f.read()
            if all(
                not any(cell.strip() for cell in line.split(","))
                for line in content.splitlines()
            ):
                f.seek(0)
                f.truncate()


def _safe_str(val: Any) -> str:
    try:
        return "" if val is None else str(val)
    except Exception:
        return ""


def _scan_defined_names_for_ref(
    wb: "openpyxl.Workbook",
) -> List[Dict[str, str]]:
    errors: List[Dict[str, str]] = []
    try:
        for dn in getattr(wb, "defined_names", []) or []:
            # dn.attr_text holds the textual formula/range
            txt = _safe_str(getattr(dn, "attr_text", None))
            if "#REF!" in txt:
                errors.append(
                    {
                        "sheet": "<defined-name>",
                        "cell": _safe_str(getattr(dn, "name", "<unnamed>")),
                        "formula": txt,
                    }
                )
    except Exception:
        # Best-effort; ignore parsing issues
        pass
    return errors


def _scan_data_validations_for_ref(ws: Worksheet) -> List[Dict[str, str]]:
    errors: List[Dict[str, str]] = []
    try:
        data_validations = getattr(ws, "data_validations", None)
        if data_validations is not None and hasattr(data_validations, "dataValidation"):
            dv_list = data_validations.dataValidation or []
        else:
            dv_list = []
        for dv in dv_list:
            # Check formula1/formula2 and sqref strings
            for attr in ("formula1", "formula2", "sqref"):
                txt = _safe_str(getattr(dv, attr, None))
                if "#REF!" in txt:
                    errors.append(
                        {
                            "sheet": ws.title,
                            "cell": _safe_str(attr),
                            "formula": txt,
                        }
                    )
    except Exception:
        pass
    return errors


def _scan_zip_for_ref_tokens(xlsx_path: Path) -> List[Dict[str, str]]:
    """
    Scan the underlying XLSX/XLSM/XLTM zip parts for literal '#REF!' tokens.
    This catches references present in XML (charts, conditional formatting,
    defined names, etc.) that may not surface via openpyxl objects.
    """
    findings: List[Dict[str, str]] = []
    try:
        with zipfile.ZipFile(xlsx_path, "r") as zf:
            for name in zf.namelist():
                # Only inspect Excel XML parts
                if not name.startswith("xl/") or not name.endswith(".xml"):
                    continue
                try:
                    data = zf.read(name)
                except Exception:
                    continue
                if b"#REF!" in data:
                    findings.append(
                        {
                            "sheet": "<xml>",
                            "cell": name,
                            "formula": "#REF! found in XML part",
                        }
                    )
    except Exception:
        # If the file isn't a zip (like xlsb) or read error, ignore; other
        # checks will handle detection.
        pass
    return findings


def check_ref_errors(file_path: Path):
    """Check for #REF! errors in Excel file."""
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(file_path, data_only=False, keep_links=False)
        ref_errors: List[Dict[str, str]] = []

        # Check each worksheet
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Skip if this is not a worksheet (e.g., chartsheet)
            if not isinstance(sheet, Worksheet):
                continue

            # Check all cells in the sheet
            for row in sheet.iter_rows():
                for cell in row:
                    # Check both the cell value and formula
                    cell_value = cell.value
                    has_ref_error = False

                    # Check if cell has a value and contains #REF!
                    if cell_value and isinstance(cell_value, str):
                        if "#REF!" in cell_value:
                            has_ref_error = True

                    # For formula cells, also check the formula itself
                    if cell.data_type == "f":
                        raw = (
                            cell_value.text
                            if isinstance(cell_value, ArrayFormula)
                            else cell_value
                        )
                        if "#REF!" in _safe_str(raw):
                            has_ref_error = True

                    if has_ref_error:
                        # Format the formula for display
                        if isinstance(cell_value, ArrayFormula):
                            formula_display = _safe_str(cell_value.text)
                        else:
                            formula_display = _safe_str(cell_value)

                        ref_errors.append(
                            {
                                "sheet": sheet_name,
                                "cell": cell.coordinate,
                                "formula": formula_display,
                            }
                        )

            # Check data validations on the sheet
            ref_errors.extend(_scan_data_validations_for_ref(sheet))

        # Check defined names at workbook level
        ref_errors.extend(_scan_defined_names_for_ref(workbook))

        # Fallback: scan raw XML parts for literal '#REF!'
        ref_errors.extend(_scan_zip_for_ref_tokens(file_path))

        workbook.close()
        return ref_errors

    except InvalidFileException as e:
        print(f"ERROR: {file_path} is not a valid Excel file: {e}")
        return None
    except Exception as e:
        print(f"ERROR: Error reading {file_path}: {str(e)}")
        return None


def delete_existing_exploded(base_path: str = "exploded"):
    """Delete existing exploded directory contents."""
    exploded_path = Path(base_path)
    if exploded_path.exists():
        for child in exploded_path.glob("*"):
            if child.is_file():
                child.unlink()
            elif child.is_dir():
                for subchild in child.glob("**/*"):
                    if subchild.is_file():
                        subchild.unlink()
                try:
                    child.rmdir()
                except OSError:
                    pass  # Directory not empty, skip


def main():
    """Main function to check Excel files for #REF! errors."""

    # Get Excel file from environment variable or command line
    excel_file = os.environ.get("EXCEL_FILE")
    export_sheets = os.environ.get("EXPORT_SHEETS", "false").lower() == "true"

    if not excel_file and len(sys.argv) > 1:
        excel_file = sys.argv[1]

    if not excel_file:
        print(
            "ERROR: No Excel file specified. Set EXCEL_FILE environment "
            "variable or pass as argument."
        )
        sys.exit(1)

    file_path = Path(excel_file)

    supported_extensions = {".xlsx", ".xltm", ".xlsm", ".xltx", ".xlsb"}
    if file_path.suffix.lower() not in supported_extensions:
        print(f"ERROR: Unsupported file type: {file_path.suffix}")
        sys.exit(1)

    if not file_path.exists():
        print(f"ERROR: File not found: {file_path}")
        sys.exit(1)

    print(f"[CHECK] Checking {file_path.name} for #REF! errors...")

    # Add debug flag
    debug = os.environ.get("DEBUG_EXCEL", "false").lower() == "true"

    if debug:
        print(f"DEBUG: File path: {file_path}")
        print(f"DEBUG: File exists: {file_path.exists()}")
        file_size = file_path.stat().st_size if file_path.exists() else "N/A"
        print(f"DEBUG: File size: {file_size}")
        print(f"DEBUG: EXCEL_FILE env var = {os.environ.get('EXCEL_FILE')}")
        print(f"DEBUG: sys.argv = {sys.argv}")

    # Check for #REF! errors
    ref_errors = check_ref_errors(file_path)

    if debug:
        print(f"Debug: Found {len(ref_errors) if ref_errors else 0} errors")

    if ref_errors is None:
        sys.exit(1)
    elif len(ref_errors) > 0:
        print(f"ERROR: Found {len(ref_errors)} #REF! errors:")
        for error in ref_errors:
            sheet = error["sheet"]
            cell = error["cell"]
            formula = error["formula"]
            print(f"  - Sheet '{sheet}', Cell {cell}: {formula}")
        print("\nERROR: Please fix broken references before merging.")
        sys.exit(1)
    else:
        print("SUCCESS: No #REF! errors found.")

    # Export sheets if requested
    if export_sheets:
        print("[EXPORT] Exporting sheets with formulas...")
        delete_existing_exploded()

        exploded_root = Path("exploded") / file_path.stem
        sheets_dir = exploded_root / "sheets"

        try:
            export_sheets_with_formulas(file_path, sheets_dir)
            print(f"SUCCESS: Sheets exported to: {sheets_dir}")
        except Exception as e:
            print(f"WARNING: Failed to export sheets: {str(e)}")

    print("SUCCESS: Excel validation completed successfully.")
    sys.exit(0)


if __name__ == "__main__":
    main()
